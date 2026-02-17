import mysql.connector
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from dotenv import load_dotenv

load_dotenv()
user_name = os.getenv("DB_USER")
password = os.getenv("DB_PASSWORD")
db_name = os.getenv("DB_NAME")
db_host = os.getenv("DB_HOST")

def crear_excel_cuota(month,year):
    conn = mysql.connector.connect(
        host = db_host,
        user = user_name,
        password = password,
        database = db_name
    )

    cursor = conn.cursor(buffered=True)
    cursor.execute("CALL cuotas_x_mes(%s,%s)",(month,year))
    excel_data = cursor.fetchall()

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    file_path = "Cuotas " + meses[month-1] + ".xlsx"

    wb = Workbook()
    ws = wb.active
    for i in excel_data:
        ws.append(i)
    wb.save(file_path)

    #Ajustar Tamaños de celdas
    wb = load_workbook(file_path)
    ws = wb.active

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 56

    # ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
    for row in ws["C"]:
        row.alignment = Alignment(horizontal="center")
    wb.save(file_path)

    return file_path
# Crear Grafico (Columnas Apiladas Horizontal?)