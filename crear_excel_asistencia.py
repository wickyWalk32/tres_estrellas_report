
from openpyxl.utils import get_column_letter
from openpyxl import  load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from dotenv import load_dotenv
import os
import psycopg2

load_dotenv()

user_name = os.getenv("DB_USER")
password = os.getenv("DB_PASSWORD")
db_name = os.getenv("DB_NAME")
db_host = os.getenv("DB_HOST")

def get_all_asistencias():
    conn = psycopg2.connect(
        #host="localhost",        # Database server address
       #port=port_number,    # specify only if default 3306 port is not the one being used
        host="localhost",
        user=user_name,
        password=password,
        database=db_name
    )
    cursor = conn.cursor() # Para ejecutar SQL queries


    cursor.execute("select * from asistencia_pivot()")
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    headers = [col[0] for col in cursor.description] # ['jugador', fecha1,fecha2,...]

    excel_data = [list(headers)] + [list(x) for x in data]
    file_path = "Asistencia.xlsx"
    return file_path


def crear_archivo(excel_data,path_file):
    wb = Workbook()
    ws = wb.active

    for row in excel_data:
        ws.append(row)
    wb.save(path_file)
    # generar_grafico(excel_data,file_path)


    # Cargar Excel existente (previamente creado)
    wb = load_workbook(path_file)
    ws = wb.active
    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1)):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = 20 if i == 0 else 12
    wb.save(path_file)
    print("Archivo Generado")


def generar_grafico(excel_data,path_file: str):
    wb = load_workbook(path_file)
    ws = wb.active
# Grafico de Columnas Apiladas
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Asistencia"
    chart.grouping = "stacked"


    chart.width = 2 * len(excel_data) / 3    # default 15
    chart.height = 25   # default 7.5
    chart.layout = Layout(
        ManualLayout( x=0.10, y=0.15, h=0.8, w=0.8, xMode="edge", yMode="edge",)
    )
    chart.legend.position = "b"  # bottom

    #Forzar a que aparezcan los ejes (asistencias y nombres)
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    chart.overlap = 100
    chart.gapWidth = 85
    chart.y_axis.majorUnit = 1


    data_ref = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    excel_chart_position = "A" + str(len(excel_data)+2)
    ws.add_chart(chart, excel_chart_position)

    wb.save(path_file)
    print("Grafico Generado")

# get_all_asistencias()
# crear_archivo(excel_data,file_path)
# generar_grafico(excel_data, file_path)

#fecha = fechas[i][0].strftime('%d-%m-%Y')

def crear_excel_asistencia(month,year):
    # Coneccion para obtener asistencias x mes
    conn = psycopg2.connect(
        host=db_host,        # Database server address
       #port=port_number,    # specify only if default 3306 port is not the one being used
        user=user_name,
        password=password,
        database=db_name
    )
    cursor2 = conn.cursor()
    cursor2.execute("select * from asistencia_pivot_by_date(%s, %s)", (month, year))
    data = cursor2.fetchall()

    cursor2.close()
    conn.close()

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    file_path = "Asistencia " + meses[month-1] + ".xlsx"

    players_attendance = data[0][0] if data and data[0] else []

    if(players_attendance == []): return []
    attendance_sorted = sorted(players_attendance, key=lambda x: x['id_jugador'])

    date_keys = sorted(
        k for k in attendance_sorted[0]
        if k not in ('jugador', 'id_jugador')
    )

    rows = [ [ p['jugador']] + [p[d] for d in date_keys] for p in attendance_sorted ]

    header = ['Jugador'] + date_keys
    excel_data = [header] + rows

    crear_archivo(excel_data,file_path)
    generar_grafico(excel_data,file_path)

    return file_path